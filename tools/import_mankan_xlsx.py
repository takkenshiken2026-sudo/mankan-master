#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""マン管マスター用: マンション管理士過去問 Excel → data/past_questions.csv 変換。

使い方:
    python3 tools/import_mankan_xlsx.py \
        --src "/Users/otedaiki/Desktop/マンション管理士過去問.xlsx" \
        [--dst data/past_questions.csv]

Excel スキーマ (1シート、ヘッダー1行):
    年度 / 問番号 / 問題文 / 問題タイプ / 正解 / 免除対象 /
    選択肢1 / 選択肢2 / 選択肢3 / 選択肢4 / 分野・根拠法令 / 解説

このスクリプトはサイト固有なので tools/template_site_only.paths で除外されます。
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("error: openpyxl が必要です。 python3 -m pip install --user openpyxl", file=sys.stderr)
    sys.exit(1)


ROOT = Path(__file__).resolve().parents[1]

# 和暦→西暦（年度）
WAREKI_YEAR: dict[str, int] = {
    "令和元年度": 2019,
    "令和2年度": 2020,
    "令和3年度": 2021,
    "令和4年度": 2022,
    "令和5年度": 2023,
    "令和6年度": 2024,
    "令和7年度": 2025,
}

# Excel の分野表記 → site-config.json の fields[].aliases に登録済みの正規名
CATEGORY_CANONICAL: dict[str, str] = {
    "区分所有法": "区分所有法",
    "区分所有法（判例）": "区分所有法（判例）",
    "民法": "民法",
    "民法（判例）": "民法（判例）",
    "借地借家法": "借地借家法",
    "標準管理規約": "標準管理規約",
    "標準管理規約（単棟型）": "標準管理規約（単棟型）",
    "標準管理規約（団地型）": "標準管理規約（団地型）",
    "標準管理規約（複合用途型）": "標準管理規約（複合用途型）",
    "マンション管理適正化法": "マンション管理適正化法",
    "適正化基本方針": "適正化基本方針",
    "マンション建替え等円滑化法": "マンション建替え等円滑化法",
    "建替え円滑化法": "建替え円滑化法",
    "被災区分所有法": "被災区分所有法",
    "被災マンション法": "被災マンション法",
    "不動産登記法": "不動産登記法",
    "民事執行法": "民事執行法",
    "都市計画法": "都市計画法",
    "建築基準法": "建築基準法",
    "水道法": "水道法",
    "消防法": "消防法",
    "警備業法": "警備業法",
    "防犯指針": "防犯指針",
    "建物・設備": "建物・設備",
    "長期修繕計画": "長期修繕計画",
    "会計": "会計",
}

ANNOTATION_RE = re.compile(r"（[\u203b※][^）]*）")
SEGMENT_SPLIT_RE = re.compile(r"、")


def clean_text(value) -> str:
    if value is None:
        return ""
    s = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    return s


def parse_category(raw: str) -> tuple[str, list[str], dict]:
    """分野・根拠法令セルを正規化。

    戻り値: (canonical_category, all_segments_canonical, special_flags)
    special_flags: {"invalidated": bool, "multiple_correct": str|None, "note": str}
    """
    flags: dict = {"invalidated": False, "multiple_correct": None, "note": ""}
    annotations = ANNOTATION_RE.findall(raw)
    for ann in annotations:
        if "正解なし" in ann:
            flags["invalidated"] = True
            flags["note"] = (flags["note"] + " / " + ann.strip("（）")).strip(" /")
        elif "正解として取り扱う" in ann or "又は" in ann:
            flags["invalidated"] = True
            m = re.search(r"(\d+(?:又は\d+)+)", ann)
            if m:
                flags["multiple_correct"] = m.group(1)
            flags["note"] = (flags["note"] + " / " + ann.strip("（）")).strip(" /")
    stripped = ANNOTATION_RE.sub("", raw).strip().strip("・、, ")
    segments_raw = [s.strip().strip("・、, ") for s in SEGMENT_SPLIT_RE.split(stripped) if s.strip().strip("・、, ")]

    canonical_segments: list[str] = []
    for seg in segments_raw:
        if seg in CATEGORY_CANONICAL:
            canonical_segments.append(CATEGORY_CANONICAL[seg])
            continue
        base = re.sub(r"（[^）]*）", "", seg).strip().strip("・、, ")
        if base in CATEGORY_CANONICAL:
            canonical_segments.append(CATEGORY_CANONICAL[base])
        else:
            canonical_segments.append(seg)

    primary = canonical_segments[0] if canonical_segments else ""
    return primary, canonical_segments, flags


def to_int(value) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(str(value).strip())
    except ValueError:
        return None


MULTI_CORRECT_RE = re.compile(r"\d+\s*(?:[、又は・,]+\s*\d+)+")


def parse_correct(raw, flags: dict) -> int | None:
    """正解セルを解釈。0 や '1、3' のような複数正解扱いは没問フラグを立てる。"""
    if raw is None or raw == "":
        return None
    s = str(raw).strip()
    if MULTI_CORRECT_RE.fullmatch(s) or "、" in s or "又は" in s:
        nums = re.findall(r"\d+", s)
        flags["invalidated"] = True
        if not flags["note"]:
            flags["note"] = f"複数正解扱い（{s}）"
        if nums:
            n = int(nums[0])
            return n if 1 <= n <= 4 else None
        return None
    if s.isdigit():
        n = int(s)
        if n == 0:
            flags["invalidated"] = True
            if not flags["note"]:
                flags["note"] = "正解なし扱い（Excel の正解=0）"
            return None
        if 1 <= n <= 4:
            return n
        flags["invalidated"] = True
        if not flags["note"]:
            flags["note"] = f"想定外の正解値: {s}"
        return None
    return None


CSV_COLUMNS = [
    "exam_year", "exam_wareki", "question_no", "type", "category", "tags",
    "stem", "preamble", "statement_a", "statement_b", "statement_c", "statement_d",
    "choice_1", "choice_2", "choice_3", "choice_4",
    "correct", "is_exempt", "is_invalidated", "note",
    "explanation", "explanation_summary", "explanation_correct",
    "explanation_choices", "explanation_point", "related_links",
]


def convert(src: Path, dst: Path) -> dict:
    wb = load_workbook(src, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    raw_rows = list(ws.iter_rows(min_row=2, values_only=True))

    out_rows: list[dict] = []
    unknown_categories: set[str] = set()
    skipped: list[str] = []

    for raw in raw_rows:
        if raw is None or all(c is None or c == "" for c in raw):
            continue
        wareki = clean_text(raw[0])
        qno = to_int(raw[1])
        stem = clean_text(raw[2])
        qtype_label = clean_text(raw[3])
        correct_raw = raw[4]
        exempt_raw = clean_text(raw[5])
        choice_1 = clean_text(raw[6])
        choice_2 = clean_text(raw[7])
        choice_3 = clean_text(raw[8])
        choice_4 = clean_text(raw[9])
        category_raw = clean_text(raw[10])
        explanation = clean_text(raw[11])

        if not wareki or qno is None:
            skipped.append(f"{wareki!r} 問{raw[1]!r}: 年度または問番号欠落")
            continue
        year = WAREKI_YEAR.get(wareki)
        if year is None:
            skipped.append(f"{wareki}: 西暦マッピング未登録")
            continue

        primary, segments, flags = parse_category(category_raw)
        for seg in segments:
            if seg not in CATEGORY_CANONICAL.values():
                unknown_categories.add(seg)

        correct = parse_correct(correct_raw, flags)
        if flags["invalidated"] and flags["multiple_correct"] and correct is None:
            first = flags["multiple_correct"].split("又は")[0].strip()
            correct = to_int(first)
        is_invalidated = "TRUE" if flags["invalidated"] else "FALSE"

        is_exempt = "TRUE" if exempt_raw in {"○", "〇"} else "FALSE"

        tags: list[str] = []
        if qtype_label:
            tags.append(qtype_label)
        for seg in segments:
            if seg and seg not in tags:
                tags.append(seg)
        if exempt_raw in {"○", "〇"}:
            tags.append("5問免除")

        note = flags["note"]

        out_rows.append({
            "exam_year": year,
            "exam_wareki": wareki,
            "question_no": qno,
            "type": "single",
            "category": primary,
            "tags": ";".join(tags),
            "stem": stem,
            "preamble": "",
            "statement_a": "",
            "statement_b": "",
            "statement_c": "",
            "statement_d": "",
            "choice_1": choice_1,
            "choice_2": choice_2,
            "choice_3": choice_3,
            "choice_4": choice_4,
            "correct": "" if (flags["invalidated"] and correct is None) else (str(correct) if correct is not None else ""),
            "is_exempt": is_exempt,
            "is_invalidated": is_invalidated,
            "note": note,
            "explanation": explanation,
            "explanation_summary": "",
            "explanation_correct": "",
            "explanation_choices": "",
            "explanation_point": "",
            "related_links": "",
        })

    out_rows.sort(key=lambda r: (r["exam_year"], r["question_no"]))

    dst.parent.mkdir(parents=True, exist_ok=True)
    with dst.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS, quoting=csv.QUOTE_MINIMAL)
        writer.writeheader()
        for row in out_rows:
            writer.writerow(row)

    return {
        "written": len(out_rows),
        "unknown_categories": sorted(unknown_categories),
        "skipped": skipped,
    }


def main() -> int:
    ap = argparse.ArgumentParser(description="マン管 Excel → past_questions.csv")
    ap.add_argument("--src", required=True, type=Path)
    ap.add_argument("--dst", type=Path, default=ROOT / "data" / "past_questions.csv")
    args = ap.parse_args()
    result = convert(args.src.resolve(), args.dst.resolve())
    print(f"Wrote {result['written']} rows -> {args.dst}")
    if result["unknown_categories"]:
        print("\n[WARN] site-config の aliases に未登録の正規化結果:")
        for c in result["unknown_categories"]:
            print(f"  - {c}")
    if result["skipped"]:
        print("\n[WARN] スキップした行:")
        for s in result["skipped"]:
            print(f"  - {s}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
