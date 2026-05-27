#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""マン管マスター用: 一問一答 Excel → data/ichimon_questions.csv 変換。

使い方:
    python3 tools/import_mankan_ichimon_xlsx.py \
        --src "/Users/otedaiki/Desktop/マンション管理士過去問_一問一答500問.xlsx" \
        [--dst data/ichimon_questions.csv]

Excel スキーマ（シート「一問一答」、ヘッダー1行）:
    年度 / 元問番号 / 一問一答番号 / 元選択肢 / 問題文 / 正誤 / 分野・根拠法令 / 解説

ichimon_questions.csv の id 形式: YYYY-問番号-枝番（例: 2024-01-1）
  → 静的 URL: q/ichimon/y2024/i01-1/index.html

このスクリプトはサイト固有なので tools/template_site_only.paths で除外されます。
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("error: openpyxl が必要です。 python3 -m pip install --user openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from tools.site_config import category_to_field_map  # noqa: E402

ANNOTATION_RE = re.compile(r"（[\u203b※][^）]*）")
WAREKI_RE = re.compile(r"令和(\d+)年度")
CATEGORY_MAP = category_to_field_map()

CSV_COLUMNS = [
    "id",
    "question",
    "answer",
    "explanation",
    "explanation_summary",
    "explanation_correct",
    "explanation_opposite",
    "explanation_point",
    "category",
    "tags",
    "source",
    "note",
]


def clean_text(value) -> str:
    if value is None:
        return ""
    s = str(value).replace("\r\n", "\n").replace("\r", "\n")
    s = re.sub(r"[ \t\u3000]+", " ", s)
    s = re.sub(r"\n+", " ", s)
    return s.strip()


def to_int(value) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(str(value).strip())
    except ValueError:
        return None


def wareki_to_year(wareki: str) -> int:
    s = clean_text(wareki)
    if "令和元" in s:
        return 2019
    m = WAREKI_RE.search(s)
    if m:
        return 2018 + int(m.group(1))
    raise ValueError(f"未対応の年度表記: {wareki!r}")


def parse_category(raw: str) -> tuple[str, list[str]]:
    """primary category（site-config 登録名）と tags 用セグメント一覧。"""
    stripped = ANNOTATION_RE.sub("", raw).strip().strip("・、, ")
    s = stripped.replace("・（", "、（")
    segments: list[str] = []
    for seg in re.split(r"[、,]", s):
        seg = seg.strip().strip("・、, ")
        if not seg:
            continue
        # 「（複合用途型）」だけ残るケースを補完
        if seg.startswith("（") and segments:
            base = segments[-1]
            if "（" in base:
                prefix = base.split("（", 1)[0]
                inner = seg.strip("（）")
                seg = f"{prefix}（{inner}）"
            else:
                continue
        segments.append(seg)
    primary = segments[0] if segments else stripped
    return primary, segments


def normalize_answer(raw) -> str:
    s = clean_text(raw)
    if s in ("○", "〇"):
        return "○"
    if s in ("×", "✕", "╳"):
        return "×"
    raise ValueError(f"想定外の正誤: {raw!r}")


def first_sentence(text: str, limit: int = 120) -> str:
    if not text:
        return ""
    parts = re.split(r"(?<=[。．])", text, maxsplit=1)
    head = parts[0].strip()
    if len(head) > limit:
        return head[: limit - 1] + "…"
    return head


def enrich_explanation_fields(
    explanation: str, answer: str, question: str, category: str
) -> dict[str, str]:
    exp = clean_text(explanation) or "（解説は未入力です。）"
    is_true = answer == "○"
    summary = first_sentence(exp)
    if not summary:
        summary = "この記述は正しい内容です。" if is_true else "この記述は誤りです。"

    if is_true:
        correct = exp
        opposite = (
            f"× を選ぶ場合は、条文の例外や限定語（「当然に」「全員」「必ず」など）を"
            f"読み落としている可能性があります。{category} の関連条文を確認してください。"
        )
    else:
        correct = exp
        opposite = (
            f"○ を選ぶ場合は、一見正しそうな一般論で判断している可能性があります。"
            f"「{first_sentence(question, 40)}」のような限定語に注意してください。"
        )

    point = (
        f"{category} の条文・判例を用語解説で確認し、"
        f"同分野の過去問・実践演習で解き直すと定着しやすくなります。"
    )
    return {
        "explanation": exp,
        "explanation_summary": summary,
        "explanation_correct": correct,
        "explanation_opposite": opposite,
        "explanation_point": point,
    }


def make_id(year: int, qno: int, branch: int) -> str:
    return f"{year}-{qno:02d}-{branch}"


def convert_row(cells: tuple) -> dict[str, str]:
    wareki, orig_qno_raw, branch_raw, choice_raw, question_raw, answer_raw, cat_raw, exp_raw = cells[:8]

    year = wareki_to_year(str(wareki or ""))
    orig_qno = to_int(orig_qno_raw)
    branch = to_int(branch_raw)
    choice = to_int(choice_raw)
    if orig_qno is None or branch is None:
        raise ValueError(f"問番号/枝番が不正: {orig_qno_raw!r}, {branch_raw!r}")

    question = clean_text(question_raw)
    if not question:
        raise ValueError("問題文が空です")

    answer = normalize_answer(answer_raw)
    primary_cat, cat_segments = parse_category(str(cat_raw or ""))
    if primary_cat not in CATEGORY_MAP:
        raise ValueError(f"未登録 category: {primary_cat!r} (raw={cat_raw!r})")

    tags = ";".join(dict.fromkeys(cat_segments))  # 重複除去・順序保持
    source = f"{clean_text(wareki)} 元問{orig_qno} 選択肢{choice or branch}"
    exp_fields = enrich_explanation_fields(
        str(exp_raw or ""), answer, question, primary_cat
    )

    return {
        "id": make_id(year, orig_qno, branch),
        "question": question,
        "answer": answer,
        "category": primary_cat,
        "tags": tags,
        "source": source,
        "note": "",
        **exp_fields,
    }


def import_xlsx(src: Path, dst: Path) -> int:
    wb = load_workbook(src, read_only=True, data_only=True)
    if "一問一答" not in wb.sheetnames:
        raise ValueError(f"シート「一問一答」がありません: {wb.sheetnames}")
    ws = wb["一問一答"]

    rows: list[dict[str, str]] = []
    seen_ids: set[str] = set()
    for line_no, cells in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(c is not None and str(c).strip() for c in cells):
            continue
        try:
            row = convert_row(cells)
        except ValueError as exc:
            raise ValueError(f"行 {line_no}: {exc}") from exc
        rid = row["id"]
        if rid in seen_ids:
            raise ValueError(f"行 {line_no}: id 重複 {rid}")
        seen_ids.add(rid)
        rows.append(row)

    wb.close()

    dst.parent.mkdir(parents=True, exist_ok=True)
    with dst.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS, quoting=csv.QUOTE_MINIMAL)
        writer.writeheader()
        writer.writerows(rows)

    print(f"Wrote {dst} ({len(rows)} rows)")
    from collections import Counter

    print("  年度分布:")
    for y, n in sorted(Counter(r["id"].split("-")[0] for r in rows).items()):
        print(f"    {y}: {n}")
    print("  正誤分布:")
    for a, n in Counter(r["answer"] for r in rows).items():
        print(f"    {a}: {n}")
    return len(rows)


def main() -> int:
    parser = argparse.ArgumentParser(description="マン管 一問一答 Excel → ichimon_questions.csv")
    parser.add_argument(
        "--src",
        type=Path,
        default=Path("/Users/otedaiki/Desktop/マンション管理士過去問_一問一答500問.xlsx"),
    )
    parser.add_argument("--dst", type=Path, default=ROOT / "data" / "ichimon_questions.csv")
    args = parser.parse_args()

    if not args.src.is_file():
        print(f"error: ファイルがありません: {args.src}", file=sys.stderr)
        return 1

    try:
        n = import_xlsx(args.src, args.dst)
    except ValueError as exc:
        print(f"error: {exc}", file=sys.stderr)
        return 1

    print(f"完了: {n} 問を {args.dst} に書き出しました")
    print("次: python3 tools/build_all.py")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
