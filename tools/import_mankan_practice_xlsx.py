#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""マン管マスター用: 実践演習 Excel → data/practice_questions.csv 変換。

使い方:
    python3 tools/import_mankan_practice_xlsx.py \
        --src "/Users/otedaiki/Desktop/マンション管理士オリジナル実践演習500問.xlsx" \
        [--dst data/practice_questions.csv]

Excel スキーマ（1シート、ヘッダー1行 / 過去問と同じ列順）:
    年度 / 問番号 / 問題文 / 問題タイプ / 正解 / 免除対象 /
    選択肢1 / 選択肢2 / 選択肢3 / 選択肢4 / 分野・根拠法令 / 解説

practice_questions.csv のスキーマ:
    question_no, type, category, tags, stem, preamble, statement_a..d,
    choice_1..4, correct, explanation,
    explanation_summary, explanation_correct, explanation_choices, explanation_point

このスクリプトはサイト固有なので tools/template_site_only.paths で除外されます。
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("error: openpyxl が必要です。 python3 -m pip install --user openpyxl", file=sys.stderr)
    sys.exit(1)


ROOT = Path(__file__).resolve().parents[1]

# 過去問コンバータと同じ canonical 表（site-config の aliases に登録済み）
CATEGORY_CANONICAL: dict[str, str] = {
    "区分所有法": "区分所有法",
    "区分所有法（判例）": "区分所有法（判例）",
    "民法": "民法",
    "民法（判例）": "民法（判例）",
    "借地借家法": "借地借家法",
    "標準管理規約": "標準管理規約",
    "標準管理規約（単棟型）": "標準管理規約（単棟型）",
    "標準管理規約（団地型）": "標準管理規約（団地型）",
    "標準管理規約（複合用途型）": "標準管理規約（複合用途型）",
    "標準管理委託契約書": "標準管理委託契約書",
    "マンション管理適正化法": "マンション管理適正化法",
    "適正化基本方針": "適正化基本方針",
    "マンション建替え等円滑化法": "マンション建替え等円滑化法",
    "建替え円滑化法": "建替え円滑化法",
    "被災区分所有法": "被災区分所有法",
    "被災マンション法": "被災マンション法",
    "不動産登記法": "不動産登記法",
    "民事執行法": "民事執行法",
    "都市計画法": "都市計画法",
    "建築基準法": "建築基準法",
    "建築法令": "建築法令",
    "建築物省エネ法": "建築物省エネ法",
    "景観法": "景観法",
    "水道法": "水道法",
    "消防法": "消防法",
    "浄化槽法": "浄化槽法",
    "警備業法": "警備業法",
    "防犯指針": "防犯指針",
    "個人情報保護": "個人情報保護",
    "建物・設備": "建物・設備",
    "長期修繕計画": "長期修繕計画",
    "耐震改修": "耐震改修",
    "建築物石綿含有建材調査": "建築物石綿含有建材調査",
    "防災": "防災",
    "会計": "会計",
    "税務": "税務",
    "保険": "保険",
    "管理組合実務": "管理組合実務",
    "管理組合運営": "管理組合運営",
    "管理委託契約": "管理委託契約",
}

ANNOTATION_RE = re.compile(r"（[\u203b※][^）]*）")
SEGMENT_SPLIT_RE = re.compile(r"、")
MULTI_CORRECT_RE = re.compile(r"\d+\s*(?:[、又は・,]+\s*\d+)+")


def clean_text(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def to_int(value) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(str(value).strip())
    except ValueError:
        return None


def parse_category(raw: str) -> tuple[str, list[str]]:
    stripped = ANNOTATION_RE.sub("", raw).strip().strip("・、, ")
    segments_raw = [s.strip().strip("・、, ") for s in SEGMENT_SPLIT_RE.split(stripped) if s.strip().strip("・、, ")]
    canonical: list[str] = []
    for seg in segments_raw:
        if seg in CATEGORY_CANONICAL:
            canonical.append(CATEGORY_CANONICAL[seg])
            continue
        base = re.sub(r"（[^）]*）", "", seg).strip().strip("・、, ")
        canonical.append(CATEGORY_CANONICAL.get(base, seg))
    primary = canonical[0] if canonical else ""
    return primary, canonical


def parse_correct(raw) -> int | None:
    """実践は没問なし想定だが、保険として複数正解/0/範囲外は None を返す。"""
    if raw is None or raw == "":
        return None
    s = str(raw).strip()
    if MULTI_CORRECT_RE.fullmatch(s) or "、" in s or "又は" in s:
        nums = re.findall(r"\d+", s)
        if nums:
            n = int(nums[0])
            return n if 1 <= n <= 4 else None
        return None
    if s.isdigit():
        n = int(s)
        return n if 1 <= n <= 4 else None
    return None


CSV_COLUMNS = [
    "question_no", "type", "category", "tags", "stem",
    "preamble", "statement_a", "statement_b", "statement_c", "statement_d",
    "choice_1", "choice_2", "choice_3", "choice_4",
    "correct", "explanation",
    "explanation_summary", "explanation_correct", "explanation_choices", "explanation_point",
]


def convert(src: Path, dst: Path) -> dict:
    wb = load_workbook(src, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    raw_rows = list(ws.iter_rows(min_row=2, values_only=True))

    out_rows: list[dict] = []
    unknown_categories: set[str] = set()
    skipped: list[str] = []
    seen_qno: set[int] = set()
    canonical_values = set(CATEGORY_CANONICAL.values())

    for raw in raw_rows:
        if raw is None or all(c is None or c == "" for c in raw):
            continue
        qno = to_int(raw[1])
        stem = clean_text(raw[2])
        qtype_label = clean_text(raw[3])
        correct = parse_correct(raw[4])
        exempt_raw = clean_text(raw[5])
        choice_1 = clean_text(raw[6])
        choice_2 = clean_text(raw[7])
        choice_3 = clean_text(raw[8])
        choice_4 = clean_text(raw[9])
        category_raw = clean_text(raw[10])
        explanation = clean_text(raw[11])

        if qno is None:
            skipped.append(f"問番号欠落: {raw!r}")
            continue
        if qno in seen_qno:
            skipped.append(f"問番号重複: {qno}")
            continue
        seen_qno.add(qno)
        if correct is None:
            skipped.append(f"問{qno}: 正解の解釈不可: {raw[4]!r}")

        primary, segments = parse_category(category_raw)
        for seg in segments:
            if seg not in canonical_values:
                unknown_categories.add(seg)

        tags: list[str] = []
        if qtype_label:
            tags.append(qtype_label)
        for seg in segments:
            if seg and seg not in tags:
                tags.append(seg)
        if exempt_raw in {"○", "〇"}:
            tags.append("5問免除")

        out_rows.append({
            "question_no": qno,
            "type": "single",
            "category": primary,
            "tags": ";".join(tags),
            "stem": stem,
            "preamble": "",
            "statement_a": "",
            "statement_b": "",
            "statement_c": "",
            "statement_d": "",
            "choice_1": choice_1,
            "choice_2": choice_2,
            "choice_3": choice_3,
            "choice_4": choice_4,
            "correct": str(correct) if correct is not None else "",
            "explanation": explanation,
            "explanation_summary": "",
            "explanation_correct": "",
            "explanation_choices": "",
            "explanation_point": "",
        })

    out_rows.sort(key=lambda r: r["question_no"])

    dst.parent.mkdir(parents=True, exist_ok=True)
    with dst.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS, quoting=csv.QUOTE_MINIMAL)
        writer.writeheader()
        for row in out_rows:
            writer.writerow(row)

    return {"written": len(out_rows), "unknown_categories": sorted(unknown_categories), "skipped": skipped}


def main() -> int:
    ap = argparse.ArgumentParser(description="マン管 実践演習 Excel → practice_questions.csv")
    ap.add_argument("--src", required=True, type=Path)
    ap.add_argument("--dst", type=Path, default=ROOT / "data" / "practice_questions.csv")
    args = ap.parse_args()
    result = convert(args.src.resolve(), args.dst.resolve())
    print(f"Wrote {result['written']} rows -> {args.dst}")
    if result["unknown_categories"]:
        print("\n[WARN] site-config に未登録の正規化結果:")
        for c in result["unknown_categories"]:
            print(f"  - {c}")
    if result["skipped"]:
        print("\n[WARN] スキップ/警告行:")
        for s in result["skipped"]:
            print(f"  - {s}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
